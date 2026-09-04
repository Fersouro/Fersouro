"""Exportacao dos modelos gold para arquivos que abrem direto no Excel.

O lake guarda Parquet, que e melhor para consulta e ocupa menos espaco. Mas
quem vai usar o numero muitas vezes quer abrir e olhar -- e para isso um .xlsx
pronto vale mais que qualquer explicacao de como ler Parquet.
"""

from __future__ import annotations

import csv
import datetime as dt
from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path
from typing import Any, Iterable

from .config import Settings
from .duck import connect, quote_literal
from .logging_conf import get_logger
from .storage import paths

log = get_logger(__name__)

# Limite fisico do formato xlsx: 1.048.576 linhas, menos a de cabecalho.
EXCEL_MAX_ROWS = 1_048_575

# Largura em caracteres; acima disso a coluna fica larga demais para ler.
_MAX_WIDTH = 60
_MIN_WIDTH = 10


@dataclass
class ExportResult:
    model: str
    path: Path | None
    rows: int
    status: str
    message: str | None = None

    @property
    def ok(self) -> bool:
        return self.status in ("success", "skipped")


def gold_models(settings: Settings) -> dict[str, Path]:
    """Modelos materializados na gold: nome -> diretorio."""
    root = settings.gold
    if not root.is_dir():
        return {}
    return {
        p.name: p
        for p in sorted(root.iterdir())
        if p.is_dir() and any(p.glob("*.parquet"))
    }


def _fetch(settings: Settings, directory: Path, limit: int | None):
    """Le o modelo da gold devolvendo (colunas, linhas, total_real)."""
    con = connect(settings)
    try:
        relation = f"read_parquet({quote_literal(paths.glob_parquet(directory))})"
        total = con.execute(f"SELECT count(*) FROM {relation}").fetchone()[0]
        sql = f"SELECT * FROM {relation}"
        if limit is not None and total > limit:
            sql += f" LIMIT {limit}"
        result = con.execute(sql)
        colunas = [d[0] for d in result.description]
        return colunas, result.fetchall(), total
    finally:
        con.close()


def _cell(value: Any) -> Any:
    """Converte para um tipo que o openpyxl grava como valor, nao como texto."""
    if isinstance(value, Decimal):
        # Excel nao tem decimal exato; float mantem o valor utilizavel em conta.
        return float(value)
    if isinstance(value, (dt.datetime, dt.date, int, float, str, bool)) or value is None:
        return value
    return str(value)


def _column_widths(colunas: list[str], linhas: list[tuple]) -> list[int]:
    """Largura por coluna a partir do conteudo, olhando so as primeiras linhas."""
    larguras = [len(str(c)) for c in colunas]
    for linha in linhas[:200]:
        for i, valor in enumerate(linha):
            if valor is not None:
                larguras[i] = max(larguras[i], len(str(valor)))
    return [max(_MIN_WIDTH, min(_MAX_WIDTH, w + 2)) for w in larguras]


def to_xlsx(colunas: list[str], linhas: list[tuple], destino: Path, aba: str) -> Path:
    """Grava um .xlsx com cabecalho fixo e colunas dimensionadas."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError(
            "Pacote 'openpyxl' nao instalado. Rode: pip install openpyxl"
        ) from exc

    destino.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = aba[:31] or "dados"        # o Excel corta nome de aba em 31

    ws.append(list(colunas))
    for linha in linhas:
        ws.append([_cell(v) for v in linha])

    negrito = Font(bold=True, color="FFFFFF")
    fundo = PatternFill("solid", fgColor="44546A")
    for celula in ws[1]:
        celula.font = negrito
        celula.fill = fundo
        celula.alignment = Alignment(vertical="center")

    for i, largura in enumerate(_column_widths(colunas, linhas), start=1):
        ws.column_dimensions[get_column_letter(i)].width = largura

    # Cabecalho sempre visivel e filtro pronto: e o primeiro que se faz na mao.
    ws.freeze_panes = "A2"
    if linhas:
        ws.auto_filter.ref = ws.dimensions

    wb.save(destino)
    return destino


def to_csv(colunas: list[str], linhas: Iterable[tuple], destino: Path) -> Path:
    """CSV para Excel em portugues: separador ';' e BOM.

    Sem o BOM o Excel abre como ANSI e quebra acentuacao; com ',' de separador
    ele joga a linha inteira numa celula so.
    """
    destino.parent.mkdir(parents=True, exist_ok=True)
    with destino.open("w", encoding="utf-8-sig", newline="") as arquivo:
        escritor = csv.writer(arquivo, delimiter=";")
        escritor.writerow(colunas)
        for linha in linhas:
            escritor.writerow(["" if v is None else v for v in linha])
    return destino


def export_model(
    settings: Settings,
    nome: str,
    directory: Path,
    formato: str = "xlsx",
    destino_dir: Path | None = None,
) -> ExportResult:
    """Exporta um modelo da gold para xlsx ou csv."""
    formato = formato.lower()
    pasta = destino_dir or (settings.root / "export")
    limite = EXCEL_MAX_ROWS if formato == "xlsx" else None

    try:
        colunas, linhas, total = _fetch(settings, directory, limite)
        destino = pasta / f"{nome}.{formato}"
        if formato == "xlsx":
            to_xlsx(colunas, linhas, destino, aba=nome)
        elif formato == "csv":
            to_csv(colunas, linhas, destino)
        else:
            return ExportResult(nome, None, 0, "failed", f"formato '{formato}' invalido")

        aviso = None
        if total > len(linhas):
            aviso = (
                f"{total:,} linhas nao cabem no xlsx; gravadas as primeiras "
                f"{len(linhas):,}. Use --format csv para levar tudo."
            )
            log.warning("[export.%s] %s", nome, aviso)

        log.info("[export.%s] %s linhas -> %s", nome, f"{len(linhas):,}", destino)
        return ExportResult(nome, destino, len(linhas), "success", aviso)

    except Exception as exc:  # noqa: BLE001
        log.error("[export.%s] falhou: %s", nome, exc)
        return ExportResult(nome, None, 0, "failed", f"{type(exc).__name__}: {exc}")


def export_all(
    settings: Settings,
    formato: str = "xlsx",
    apenas: list[str] | None = None,
    destino_dir: Path | None = None,
) -> list[ExportResult]:
    modelos = gold_models(settings)
    if apenas:
        querido = {m.lower() for m in apenas}
        desconhecidos = querido - set(modelos)
        if desconhecidos:
            raise ValueError(
                f"Modelo gold inexistente: {', '.join(sorted(desconhecidos))}. "
                f"Disponiveis: {', '.join(sorted(modelos)) or '(nenhum)'}"
            )
        modelos = {k: v for k, v in modelos.items() if k in querido}
    if not modelos:
        log.warning("Nenhum modelo na gold para exportar")
        return []
    return [
        export_model(settings, nome, pasta, formato, destino_dir)
        for nome, pasta in modelos.items()
    ]
