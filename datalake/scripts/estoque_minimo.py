# -*- coding: utf-8 -*-
"""
Estoque Minimo de Pecas -- Revenda 1.

Le a QUANTIDADE DISPONIVEL de verdade da PEC_ITEM_ESTOQUE que JA esta no
datalake, cruza com a lista de minimos e gera a planilha marcando em VERMELHO
o que atingiu/ficou abaixo do minimo (precisa comprar).

Nao inventa numero. Se uma peca nao aparecer na tabela, ela vai listada como
"NAO ENCONTRADA" -- nunca com valor calculado.

Roda no servidor (nao precisa de Oracle):

    python C:\\datalake\\estoque_minimo.py

Opcional -- caminho do lake e revenda:

    python C:\\datalake\\estoque_minimo.py "C:\\datalake\\lake.duckdb" 1
"""
import os
import sys
import glob
import datetime
import duckdb

TAB = "ccm__pec_item_estoque"

# lista de minimos que o Fernando passou:  codigo publico -> minimo
MINIMOS = {
    "5Z0820411E": 2,
    "05E145933": 15,
    "JZZ129620M": 15,
    "2G5941036B": 1,
    "2QB201511": 130,
    "5Q0407183K": 4,
    "2G6827517A": 2,
    "04C109479J": 50,
    "5U0809958D": 1,
    "04C905607": 60,
}


def normcode(s):
    return "".join(ch for ch in str(s).upper() if ch.isalnum())


def achar_lake(argv):
    if len(argv) > 1 and os.path.isfile(argv[1]):
        return argv[1]
    for c in [r"C:\datalake\lake.duckdb",
              r"C:\datalake\lake\lake.duckdb",
              os.path.join(os.getcwd(), "lake.duckdb")]:
        if os.path.isfile(c):
            return c
    hits = glob.glob(r"C:\datalake\**\lake.duckdb", recursive=True)
    return hits[0] if hits else None


def norm_sql(expr):
    return ("regexp_replace(upper(cast(" + expr +
            " as varchar)),'[^0-9A-Z]','','g')")


def qcol(nome):
    return '"' + nome.replace('"', '""') + '"'


def main():
    revenda = None
    lake = achar_lake(sys.argv)
    # segundo argumento pode ser a revenda
    for a in sys.argv[2:]:
        if str(a).isdigit():
            revenda = int(a)
    if revenda is None:
        revenda = 1

    if not lake:
        print("NAO ACHEI lake.duckdb. Passe o caminho como 1o argumento.")
        return
    print("Lake:", lake)
    con = duckdb.connect(lake, read_only=True)

    if not con.execute("SELECT count(*) FROM information_schema.tables "
                       "WHERE table_name = ?", [TAB]).fetchone()[0]:
        print("A tabela", TAB, "nao esta no lake. Rode a carga no servidor primeiro.")
        return

    colunas = [r[0] for r in con.execute(
        "SELECT column_name FROM information_schema.columns "
        "WHERE table_name = ? ORDER BY ordinal_position", [TAB]).fetchall()]

    # ---- coluna DISPONIVEL ----
    disp = None
    for c in colunas:
        if "dispon" in c.lower():
            disp = c
            break
    if disp is None:  # fallback: alguma qtd/saldo
        for c in colunas:
            if any(k in c.lower() for k in ("saldo", "qtd", "quant", "estoq")):
                disp = c
                break
    if disp is None:
        print("Nao achei coluna de disponivel/quantidade em", TAB)
        print("colunas:", colunas)
        return
    print("Coluna de disponivel:", disp)

    # ---- coluna de CODIGO: a que casar com mais codigos da lista ----
    alvo = list(MINIMOS.keys())
    cand = [c for c in colunas if any(k in c.lower() for k in
            ("item", "codigo", "cod", "refer", "peca", "produto"))] or colunas
    melhor, melhor_n = None, 0
    for c in cand:
        try:
            n = con.execute(
                "SELECT count(DISTINCT " + norm_sql(qcol(c)) + ") FROM " + TAB +
                " WHERE " + norm_sql(qcol(c)) + " IN " +
                "(" + ",".join("?" for _ in alvo) + ")", alvo).fetchone()[0]
            if n > melhor_n:
                melhor, melhor_n = c, n
        except Exception:
            pass
    if not melhor:
        print("Nenhuma coluna casou com os codigos da lista.")
        print("=> O codigo na tabela deve estar em formato diferente. Rode pec_estoque.py.")
        return
    codcol = melhor
    print("Coluna de codigo:", codcol, "(casou", melhor_n, "de", len(alvo), "codigos)")

    # ---- coluna de REVENDA e de DESCRICAO (se existirem) ----
    revcol = next((c for c in colunas if c.lower() in
                   ("revenda", "empresa_revenda", "cod_revenda")), None)
    desccol = next((c for c in colunas if "descr" in c.lower()), None)
    print("Coluna de revenda:", revcol, "| descricao:", desccol)
    print("Revenda filtrada:", revenda if revcol else "(tabela nao tem revenda -> total)")
    print()

    # ---- consulta o disponivel de cada codigo ----
    resultado = []
    for cod, minimo in MINIMOS.items():
        codn = normcode(cod)
        where = norm_sql(qcol(codcol)) + " = ?"
        params = [codn]
        if revcol:
            where += " AND cast(" + qcol(revcol) + " as varchar) = ?"
            params.append(str(revenda))
        selcols = "sum(cast(" + qcol(disp) + " as double)) AS disp, count(*) n"
        if desccol:
            selcols = "any_value(" + qcol(desccol) + ") AS descr, " + selcols
        row = con.execute("SELECT " + selcols + " FROM " + TAB +
                          " WHERE " + where, params).fetchone()
        if desccol:
            descr, dispv, n = row
        else:
            descr, (dispv, n) = "", row
        if n == 0:
            resultado.append((cod, descr or "", None, minimo))
        else:
            resultado.append((cod, descr or "", float(dispv or 0), minimo))

    # ---- imprime tabela ----
    print("=" * 74)
    print("ESTOQUE MINIMO -- REVENDA", revenda)
    print("=" * 74)
    print("{:<12} {:>10} {:>8} {:>8}  {}".format(
        "CODIGO", "DISPONIVEL", "MINIMO", "FALTA", "SITUACAO"))
    print("-" * 74)
    for cod, descr, dispv, minimo in resultado:
        if dispv is None:
            print("{:<12} {:>10} {:>8} {:>8}  {}".format(
                cod, "NAO ENCON", minimo, "-", "peca nao encontrada na tabela"))
        else:
            falta = max(minimo - dispv, 0)
            sit = "COMPRAR" if dispv < minimo else "ok"
            print("{:<12} {:>10.0f} {:>8} {:>8.0f}  {}".format(
                cod, dispv, minimo, falta, sit))
    print("=" * 74)

    # ---- gera planilha ----
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M")
    saida_dir = os.path.dirname(lake) or "."
    xlsx_path = os.path.join(saida_dir, "estoque_minimo_revenda%d_%s.xlsx" % (revenda, ts))
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        wb = Workbook()
        ws = wb.active
        ws.title = "Estoque Minimo R%d" % revenda
        hdr = ["Codigo", "Descricao", "Disponivel", "Minimo", "Falta comprar", "Situacao"]
        ws.append(hdr)
        head_fill = PatternFill("solid", fgColor="1F3864")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = head_fill
            cell.alignment = Alignment(horizontal="center")
        red = PatternFill("solid", fgColor="F8CBAD")
        red_font = Font(color="9C0006", bold=True)
        gray = PatternFill("solid", fgColor="F2F2F2")
        thin = Side(style="thin", color="D9D9D9")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for cod, descr, dispv, minimo in resultado:
            if dispv is None:
                ws.append([cod, descr, "NAO ENCONTRADA", minimo, "", "verificar codigo"])
                comprar = True
            else:
                falta = max(minimo - dispv, 0)
                comprar = dispv < minimo
                ws.append([cod, descr, dispv, minimo, falta if comprar else 0,
                           "COMPRAR" if comprar else "OK"])
            r = ws.max_row
            for cell in ws[r]:
                cell.border = border
                if comprar:
                    cell.fill = red
                    cell.font = red_font
                elif r % 2 == 0:
                    cell.fill = gray
        widths = [14, 40, 12, 10, 14, 18]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + i)].width = w
        ws.freeze_panes = "A2"
        wb.save(xlsx_path)
        print("\nPlanilha gerada:", xlsx_path)
    except ImportError:
        print("\n(openpyxl nao instalado -- gerando CSV)")
        import csv
        csv_path = xlsx_path.replace(".xlsx", ".csv")
        with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f, delimiter=";")
            w.writerow(["Codigo", "Descricao", "Disponivel", "Minimo",
                        "Falta comprar", "Situacao"])
            for cod, descr, dispv, minimo in resultado:
                if dispv is None:
                    w.writerow([cod, descr, "NAO ENCONTRADA", minimo, "", "verificar"])
                else:
                    falta = max(minimo - dispv, 0)
                    w.writerow([cod, descr, dispv, minimo,
                                falta if dispv < minimo else 0,
                                "COMPRAR" if dispv < minimo else "OK"])
        print("CSV gerado:", csv_path)


if __name__ == "__main__":
    main()
