# -*- coding: utf-8 -*-
"""
Gera a pagina de Estoque Minimo (HTML) e a planilha (xlsx) a partir do lake.

Le o disponivel de verdade (silver: PEC_ITEM_REVENDA.qtd_contabil, o mesmo
numero da consulta gerencial do ERP) e cruza com a lista de minimos.

Alem da foto de hoje, guarda um snapshot diario em
<DATALAKE_ROOT>/historico_estoque/AAAA-MM-DD.parquet (o ultimo carregamento
do dia manda). A pagina ganha um seletor de data que mostra o disponivel de
qualquer dia ja registrado. O historico comeca a valer a partir de agora --
o ERP so tem o saldo do momento, entao dias passados nao dao para recuperar.

Fonte dos minimos, nesta ordem:
  1. <DATALAKE_ROOT>/minimos_pecas.csv   (o que a equipe edita -- persiste)
  2. <projeto>/conf/minimos_pecas.csv    (semente; copiada para (1) na 1a vez)

Saidas (na pasta export do lake):
  - estoque_minimo.html   pagina (todas as revendas, filtro, busca, data)
  - estoque_minimo_<data_hora>.xlsx   planilha do dia (COMPRAR em vermelho)

Nao inventa numero: peca sem saldo aparece com disponivel 0; codigo que nao
existe no cadastro aparece como "nao cadastrado".

Uso:
    python gerar_estoque.py [caminho_lake.duckdb]
"""
from __future__ import annotations

import os
import sys
import glob
import json
import html
import shutil
import datetime
import duckdb


def achar_lake(argv):
    for a in argv[1:]:
        if a.lower().endswith(".duckdb") and os.path.isfile(a):
            return a
    for c in [r"C:\datalake\lake.duckdb",
              os.path.join(os.getcwd(), "lake.duckdb")]:
        if os.path.isfile(c):
            return c
    hits = glob.glob(r"C:\datalake\**\lake.duckdb", recursive=True)
    return hits[0] if hits else None


def achar_minimos(lake_dir, script_dir):
    """Prefere o CSV editavel no lake; senao usa a semente do projeto e copia."""
    estavel = os.path.join(lake_dir, "minimos_pecas.csv")
    if os.path.isfile(estavel):
        return estavel
    semente = os.path.normpath(os.path.join(script_dir, "..", "conf", "minimos_pecas.csv"))
    if os.path.isfile(semente):
        try:
            shutil.copyfile(semente, estavel)
            print("Semente de minimos copiada para", estavel, "(edite esse arquivo)")
            return estavel
        except OSError:
            return semente
    return None


def qlit(s):
    return "'" + str(s).replace("'", "''") + "'"


def _consultar(con, minimos):
    """Le o disponivel de hoje cruzando os minimos com o silver. -> lista|None."""
    sql = """
        SELECT m.codigo,
               CAST(m.revenda AS INTEGER)               AS revenda,
               CAST(m.minimo  AS DOUBLE)                AS minimo,
               trim(pie.des_item_estoque)               AS descricao,
               CAST(rev.qtd_contabil AS DOUBLE)         AS disponivel,
               (pie.item_estoque IS NOT NULL)           AS cadastrada
          FROM read_csv_auto(%s, delim=';', header=true,
                             columns={'codigo':'VARCHAR','revenda':'INTEGER','minimo':'DOUBLE'}) m
          LEFT JOIN silver.ccm__pec_item_estoque pie
            ON regexp_replace(upper(CAST(pie.item_estoque_pub AS VARCHAR)),'[^0-9A-Z]','','g')
             = regexp_replace(upper(CAST(m.codigo AS VARCHAR)),'[^0-9A-Z]','','g')
          LEFT JOIN silver.ccm__pec_item_revenda rev
            ON rev.item_estoque = pie.item_estoque
           AND rev.revenda = CAST(m.revenda AS INTEGER)
         ORDER BY m.revenda, m.codigo
    """ % qlit(minimos)
    try:
        rows = con.execute(sql).fetchall()
        cols = [d[0] for d in con.description]
    except duckdb.CatalogException as exc:
        print("Faltam tabelas no lake (rode a carga primeiro):", str(exc).split(chr(10))[0])
        return None
    dados = []
    for r in rows:
        d = dict(zip(cols, r))
        cad = bool(d["cadastrada"])
        disp = None if not cad else float(d["disponivel"] or 0)
        minimo = float(d["minimo"] or 0)
        dados.append({
            "revenda": int(d["revenda"]),
            "codigo": d["codigo"],
            "descricao": (d["descricao"] or "") if cad else "(nao cadastrada no ERP)",
            "disponivel": disp,
            "minimo": minimo,
            "comprar": (disp is None) or (disp < minimo),
        })
    return dados


def _gravar_snapshot(hist_dir, dia_iso, dados):
    """Grava a foto do dia num parquet (sobrescreve o do dia)."""
    os.makedirs(hist_dir, exist_ok=True)
    destino = os.path.join(hist_dir, dia_iso + ".parquet")
    w = duckdb.connect()
    try:
        w.execute("CREATE TABLE s(data DATE, revenda INTEGER, codigo VARCHAR, "
                  "descricao VARCHAR, disponivel DOUBLE, minimo DOUBLE, comprar BOOLEAN)")
        w.executemany("INSERT INTO s VALUES (?,?,?,?,?,?,?)",
                      [(dia_iso, d["revenda"], d["codigo"], d["descricao"],
                        d["disponivel"], d["minimo"], d["comprar"]) for d in dados])
        w.execute("COPY s TO %s (FORMAT PARQUET)" % qlit(destino))
        print("Snapshot do dia:", destino)
    finally:
        w.close()


def _ler_historico(hist_dir, max_dias=365):
    """{data_iso: [linhas]} dos ultimos dias registrados, mais novo primeiro."""
    if not os.path.isdir(hist_dir) or not glob.glob(os.path.join(hist_dir, "*.parquet")):
        return {}
    w = duckdb.connect()
    try:
        rows = w.execute(
            "SELECT CAST(data AS VARCHAR), revenda, codigo, descricao, "
            "disponivel, minimo, comprar FROM read_parquet(%s) "
            "ORDER BY data DESC, revenda, codigo"
            % qlit(os.path.join(hist_dir, "*.parquet"))).fetchall()
    finally:
        w.close()
    hist = {}
    for dia, rev, cod, desc, disp, mi, comp in rows:
        hist.setdefault(dia, []).append({
            "revenda": int(rev),
            "codigo": cod,
            "descricao": desc or "",
            "disponivel": (None if disp is None else float(disp)),
            "minimo": float(mi or 0),
            "comprar": bool(comp),
        })
    datas = sorted(hist.keys(), reverse=True)[:max_dias]
    return {d: hist[d] for d in datas}


def main():
    lake = achar_lake(sys.argv)
    if not lake:
        print("NAO achei lake.duckdb.")
        return 1
    lake_dir = os.path.dirname(lake) or "."
    script_dir = os.path.dirname(os.path.abspath(__file__))
    minimos = achar_minimos(lake_dir, script_dir)
    if not minimos:
        print("NAO achei minimos_pecas.csv (nem no lake, nem em conf/).")
        return 1
    print("Lake:", lake)
    print("Minimos:", minimos)

    con = duckdb.connect(lake, read_only=True)
    dados = _consultar(con, minimos)
    con.close()
    if dados is None:
        return 1

    hoje = datetime.date.today().isoformat()
    hist_dir = os.path.join(lake_dir, "historico_estoque")
    _gravar_snapshot(hist_dir, hoje, dados)
    historico = _ler_historico(hist_dir)
    if hoje not in historico:            # garante o dia atual mesmo sem parquet
        historico = {hoje: [dict(revenda=d["revenda"], codigo=d["codigo"],
                                 descricao=d["descricao"], disponivel=d["disponivel"],
                                 minimo=d["minimo"], comprar=d["comprar"]) for d in dados],
                     **historico}

    gerado = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")
    comprar_n = sum(1 for d in dados if d["comprar"])
    print("Linhas hoje:", len(dados), "| para comprar:", comprar_n,
          "| dias no historico:", len(historico))

    export_dir = os.path.join(lake_dir, "export")
    os.makedirs(export_dir, exist_ok=True)
    _escrever_html(export_dir, historico, gerado)
    _escrever_xlsx(export_dir, dados)
    return 0


def _fmt_data(iso):
    a, m, d = iso.split("-")
    return "%s/%s/%s" % (d, m, a)


def _escrever_html(export_dir, historico, gerado):
    datas = list(historico.keys())          # ja vem mais novo primeiro
    revendas = sorted({r["revenda"] for linhas in historico.values() for r in linhas})
    data_opts = "".join("<option value='%s'>%s</option>" % (d, _fmt_data(d)) for d in datas)
    rev_opts = "".join("<option value='%d'>Revenda %d</option>" % (r, r) for r in revendas)
    page = (_TEMPLATE
            .replace("__HIST__", json.dumps(historico, ensure_ascii=False))
            .replace("__DATA_OPTS__", data_opts)
            .replace("__REV_OPTS__", rev_opts)
            .replace("__GERADO__", html.escape(gerado)))
    caminho = os.path.join(export_dir, "estoque_minimo.html")
    with open(caminho, "w", encoding="utf-8") as f:
        f.write(page)
    print("Pagina gerada:", caminho)


def _escrever_xlsx(export_dir, dados):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("(openpyxl ausente -- so a pagina HTML)")
        return
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M")
    caminho = os.path.join(export_dir, "estoque_minimo_%s.xlsx" % ts)
    wb = Workbook(); ws = wb.active; ws.title = "Estoque Minimo"
    ws.append(["Revenda", "Codigo", "Descricao", "Disponivel", "Minimo",
               "Falta comprar", "Situacao"])
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F3864")
        c.alignment = Alignment(horizontal="center")
    red = PatternFill("solid", fgColor="F8CBAD"); rf = Font(color="9C0006", bold=True)
    for d in sorted(dados, key=lambda x: (not x["comprar"], x["revenda"], x["codigo"])):
        disp = "NAO CADASTRADO" if d["disponivel"] is None else d["disponivel"]
        falta = "" if d["disponivel"] is None else (max(d["minimo"] - d["disponivel"], 0)
                                                    if d["comprar"] else 0)
        ws.append([d["revenda"], d["codigo"], d["descricao"], disp, d["minimo"],
                   falta, "COMPRAR" if d["comprar"] else "OK"])
        if d["comprar"]:
            for c in ws[ws.max_row]:
                c.fill = red; c.font = rf
    for i, w in enumerate([9, 14, 42, 12, 10, 14, 12], 1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A2"
    wb.save(caminho)
    print("Planilha gerada:", caminho)


_TEMPLATE = r"""<!doctype html>
<html lang="pt-br">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Estoque Minimo de Pecas - Grupo Terrasul</title>
<style>
  :root { --azul:#1F3864; --vermelho:#9C0006; --vermelhobg:#F8CBAD; --linha:#e6e6e6; }
  * { box-sizing:border-box; }
  body { margin:0; font-family:Segoe UI,Roboto,Arial,sans-serif; color:#222; background:#f4f5f7; }
  header { background:var(--azul); color:#fff; padding:16px 22px; }
  header h1 { margin:0; font-size:20px; }
  header .sub { opacity:.85; font-size:13px; margin-top:4px; }
  .barra { display:flex; flex-wrap:wrap; gap:12px; align-items:center; padding:14px 22px; background:#fff; border-bottom:1px solid var(--linha); }
  .barra label { font-size:13px; color:#555; }
  select, input[type=search] { padding:8px 10px; border:1px solid #ccc; border-radius:6px; font-size:14px; }
  input[type=search] { min-width:220px; }
  .chk { display:flex; align-items:center; gap:6px; font-size:14px; }
  .cards { display:flex; gap:12px; padding:14px 22px 0; flex-wrap:wrap; }
  .card { background:#fff; border:1px solid var(--linha); border-radius:8px; padding:12px 16px; min-width:150px; }
  .card .n { font-size:26px; font-weight:700; }
  .card.alerta .n { color:var(--vermelho); }
  .card .l { font-size:12px; color:#666; text-transform:uppercase; letter-spacing:.04em; }
  .wrap { padding:14px 22px 30px; }
  .tabela-scroll { max-height:70vh; overflow:auto; border:1px solid var(--linha); border-radius:8px; background:#fff; }
  table { width:100%; border-collapse:collapse; }
  th, td { padding:9px 12px; text-align:left; font-size:14px; border-bottom:1px solid var(--linha); }
  thead th { background:#f0f1f3; cursor:pointer; user-select:none; white-space:nowrap; position:sticky; top:0; z-index:2; box-shadow:inset 0 -1px 0 var(--linha); }
  th.num, td.num { text-align:right; font-variant-numeric:tabular-nums; }
  tr.comprar td { background:var(--vermelhobg); color:var(--vermelho); font-weight:600; }
  .tag { display:inline-block; padding:2px 8px; border-radius:10px; font-size:12px; font-weight:700; }
  .tag.c { background:var(--vermelho); color:#fff; }
  .tag.ok { background:#2e7d32; color:#fff; }
  .rodape { padding:10px 22px 30px; color:#888; font-size:12px; }
  button { padding:8px 14px; border:0; background:var(--azul); color:#fff; border-radius:6px; font-size:14px; cursor:pointer; }
  .vazio { padding:30px; text-align:center; color:#888; }
</style>
</head>
<body>
<header>
  <h1>Estoque Minimo de Pecas</h1>
  <div class="sub">Grupo Terrasul &middot; disponivel do datalake &middot; pagina gerada em __GERADO__</div>
</header>

<div class="barra">
  <label>Data
    <select id="fData">__DATA_OPTS__</select>
  </label>
  <label>Revenda
    <select id="fRev"><option value="">Todas</option>__REV_OPTS__</select>
  </label>
  <input id="fBusca" type="search" placeholder="Buscar codigo ou descricao...">
  <span class="chk"><input type="checkbox" id="fComprar"> <label for="fComprar">so o que precisa comprar</label></span>
  <button id="btnCsv">Exportar CSV</button>
</div>

<div class="cards">
  <div class="card"><div class="n" id="cTotal">0</div><div class="l">Pecas monitoradas</div></div>
  <div class="card alerta"><div class="n" id="cComprar">0</div><div class="l">Precisam comprar</div></div>
  <div class="card"><div class="n" id="cVisiveis">0</div><div class="l">Na tela</div></div>
</div>

<div class="wrap">
  <div class="tabela-scroll">
    <table>
      <thead><tr>
        <th data-k="revenda" class="num">Rev.</th>
        <th data-k="codigo">Codigo</th>
        <th data-k="descricao">Descricao</th>
        <th data-k="disponivel" class="num">Disponivel</th>
        <th data-k="minimo" class="num">Minimo</th>
        <th data-k="falta" class="num">Falta</th>
        <th data-k="comprar">Situacao</th>
      </tr></thead>
      <tbody id="corpo"></tbody>
    </table>
  </div>
  <div class="vazio" id="vazio" style="display:none">Nenhuma peca com esses filtros.</div>
</div>
<div class="rodape">Fonte: PEC_ITEM_REVENDA (QTD_CONTABIL) cruzada com a lista de minimos. Historico diario a partir do primeiro registro.</div>

<script>
var HIST = __HIST__;
var sortK = "comprar", sortDir = -1;
function dataSel(){ return document.getElementById('fData').value; }
function baseAtual(){ return (HIST[dataSel()] || []).map(function(d){
  return {revenda:d.revenda, codigo:d.codigo, descricao:d.descricao,
          disponivel:d.disponivel, minimo:d.minimo, comprar:d.comprar,
          falta:(d.disponivel===null?null:Math.max(d.minimo-d.disponivel,0))};
}); }
function fmt(v){ return v===null||v===undefined ? "-" : (Math.round(v*100)/100).toLocaleString('pt-BR'); }
function filtra(base){
  var rev = document.getElementById('fRev').value;
  var q = document.getElementById('fBusca').value.trim().toLowerCase();
  var soC = document.getElementById('fComprar').checked;
  return base.filter(function(d){
    if (rev && String(d.revenda)!==rev) return false;
    if (soC && !d.comprar) return false;
    if (q && ((d.codigo+" "+d.descricao).toLowerCase().indexOf(q)<0)) return false;
    return true;
  });
}
function ordena(a,b){
  var va=a[sortK], vb=b[sortK];
  if (typeof va==='boolean'){ va=va?1:0; vb=vb?1:0; }
  if (va===null) va=-1; if (vb===null) vb=-1;
  if (va<vb) return -sortDir; if (va>vb) return sortDir; return 0;
}
function pinta(){
  var base = baseAtual();
  var lista = filtra(base).sort(ordena);
  var tb = document.getElementById('corpo'); tb.innerHTML="";
  lista.forEach(function(d){
    var tr = document.createElement('tr'); if (d.comprar) tr.className="comprar";
    var disp = d.disponivel===null ? "n/cad" : fmt(d.disponivel);
    var sit = d.comprar ? "<span class='tag c'>COMPRAR</span>" : "<span class='tag ok'>OK</span>";
    tr.innerHTML =
      "<td class='num'>"+d.revenda+"</td>"+
      "<td>"+d.codigo+"</td>"+
      "<td>"+(d.descricao||"")+"</td>"+
      "<td class='num'>"+disp+"</td>"+
      "<td class='num'>"+fmt(d.minimo)+"</td>"+
      "<td class='num'>"+(d.falta===null?"-":fmt(d.falta))+"</td>"+
      "<td>"+sit+"</td>";
    tb.appendChild(tr);
  });
  document.getElementById('vazio').style.display = lista.length? "none":"block";
  document.getElementById('cTotal').textContent = base.length;
  document.getElementById('cComprar').textContent = base.filter(function(d){return d.comprar;}).length;
  document.getElementById('cVisiveis').textContent = lista.length;
}
document.getElementById('fData').onchange = pinta;
document.getElementById('fRev').onchange = pinta;
document.getElementById('fComprar').onchange = pinta;
document.getElementById('fBusca').oninput = pinta;
Array.prototype.forEach.call(document.querySelectorAll('thead th'), function(th){
  th.onclick = function(){
    var k = th.getAttribute('data-k');
    if (sortK===k) sortDir=-sortDir; else { sortK=k; sortDir=1; }
    pinta();
  };
});
document.getElementById('btnCsv').onclick = function(){
  var lista = filtra(baseAtual()).sort(ordena);
  var linhas = [["Data","Revenda","Codigo","Descricao","Disponivel","Minimo","Falta","Situacao"]];
  var dt = dataSel();
  lista.forEach(function(d){
    linhas.push([dt, d.revenda, d.codigo, d.descricao,
      d.disponivel===null?"n/cad":d.disponivel, d.minimo,
      d.falta===null?"":d.falta, d.comprar?"COMPRAR":"OK"]);
  });
  var csv = linhas.map(function(l){ return l.map(function(c){
    c=String(c).replace(/"/g,'""'); return /[;"\n]/.test(c)?'"'+c+'"':c; }).join(";"); }).join("\n");
  var blob = new Blob(["\ufeff"+csv], {type:"text/csv;charset=utf-8"});
  var a = document.createElement("a"); a.href=URL.createObjectURL(blob);
  a.download="estoque_minimo_"+dt+".csv"; a.click();
};
pinta();
</script>
</body>
</html>
"""


if __name__ == "__main__":
    raise SystemExit(main())
