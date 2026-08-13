"""Exportacao dos modelos gold para xlsx e csv."""

from __future__ import annotations

import datetime as dt
from decimal import Decimal

import pytest

from datalake.export import EXCEL_MAX_ROWS, _cell, _column_widths, export_all, to_csv, to_xlsx


def test_cell_converte_decimal_para_numero():
    """Decimal como texto no Excel impede somar a coluna."""
    assert _cell(Decimal("10.50")) == 10.5
    assert isinstance(_cell(Decimal("3")), float)


def test_cell_preserva_tipos_nativos():
    momento = dt.datetime(2026, 8, 13, 10, 0)
    assert _cell(momento) is momento
    assert _cell(None) is None
    assert _cell("texto") == "texto"
    assert _cell(7) == 7


def test_largura_respeita_minimo_e_maximo():
    larguras = _column_widths(["a", "descricao"], [("x", "y" * 500)])
    assert larguras[0] >= 10          # minimo legivel
    assert larguras[1] <= 60          # nao vira uma coluna gigante


def test_xlsx_grava_cabecalho_e_dados(tmp_path):
    from openpyxl import load_workbook

    destino = to_xlsx(
        ["id", "valor", "data"],
        [(1, Decimal("10.50"), dt.datetime(2026, 8, 13)), (2, None, None)],
        tmp_path / "teste.xlsx",
        aba="margem_pecas",
    )
    wb = load_workbook(destino)
    ws = wb.active
    assert ws.title == "margem_pecas"
    assert [c.value for c in ws[1]] == ["id", "valor", "data"]
    assert ws.cell(2, 2).value == 10.5          # numero, nao texto
    assert ws.freeze_panes == "A2"              # cabecalho fixo
    assert ws.auto_filter.ref is not None


def test_nome_de_aba_longo_e_cortado(tmp_path):
    from openpyxl import load_workbook

    destino = to_xlsx(["a"], [(1,)], tmp_path / "x.xlsx", aba="n" * 50)
    assert len(load_workbook(destino).active.title) == 31   # limite do Excel


def test_csv_usa_ponto_e_virgula_e_bom(tmp_path):
    """Excel em portugues: sem BOM quebra acento, com ',' junta tudo numa celula."""
    destino = to_csv(["nome", "valor"], [("Peça", 10)], tmp_path / "t.csv")
    bruto = destino.read_bytes()
    assert bruto.startswith(b"\xef\xbb\xbf")
    assert b"nome;valor" in bruto
    assert "Peça" in destino.read_text(encoding="utf-8-sig")


def test_export_all_sem_modelos_nao_falha(settings):
    assert export_all(settings) == []


def test_export_all_modelo_inexistente(settings):
    with pytest.raises(ValueError, match="inexistente"):
        export_all(settings, apenas=["nao_existe"])


def test_limite_do_excel_e_o_do_formato():
    assert EXCEL_MAX_ROWS == 1_048_575          # 1.048.576 menos o cabecalho
