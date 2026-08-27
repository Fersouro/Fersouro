"""
Gera a planilha de reposicao de estoque minimo (Revenda 1) a partir de um CSV
com o resultado da query estoque_minimo_revenda1.sql.

Uso:
    python gerar_xls_estoque_minimo.py dados/estoque_minimo_revenda1.csv saida.xlsx
    python gerar_xls_estoque_minimo.py ... --revenda 1 --fator 1.0

CSV de entrada (separador ';', UTF-8, com cabecalho):
    codigo_peca;descricao_peca;estoque_atual;estoque_minimo

A planilha sai com FORMULAS (nao valores calculados no Python), entao ao mudar
o Estoque Atual, o Estoque Minimo ou o Fator de Reposicao o Excel recalcula
sozinho o deficit, a quantidade a comprar e o status.
"""

import argparse
import csv
import datetime as dt
import os
import sys

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule

FONTE = "Arial"

AZUL = Font(name=FONTE, size=10, color="0000FF")          # entrada digitada
PRETO = Font(name=FONTE, size=10)                          # formula
CAB = Font(name=FONTE, size=10, bold=True, color="FFFFFF")
TITULO = Font(name=FONTE, size=14, bold=True)
SUB = Font(name=FONTE, size=10, italic=True, color="595959")
NEGRITO = Font(name=FONTE, size=10, bold=True)

FILL_CAB = PatternFill("solid", fgColor="1F3864")
FILL_INPUT = PatternFill("solid", fgColor="FFFF00")        # premissa a editar
FILL_TOTAL = PatternFill("solid", fgColor="D9E1F2")
FILL_CRITICO = PatternFill("solid", fgColor="FFC7CE")
FILL_ABAIXO = PatternFill("solid", fgColor="FFEB9C")
FILL_NO_MINIMO = PatternFill("solid", fgColor="DDEBF7")

FINA = Side(style="thin", color="BFBFBF")
BORDA = Border(left=FINA, right=FINA, top=FINA, bottom=FINA)

COLUNAS = [
    ("#", 5),
    ("Código da Peça", 18),
    ("Descrição da Peça", 42),
    ("Revenda", 9),
    ("Estoque Atual", 13),
    ("Estoque Mínimo", 14),
    ("Déficit", 10),
    ("Qtd. a Comprar", 14),
    ("Status", 18),
]
LINHA_CAB = 8   # linha do cabecalho da tabela
LINHA_1 = 9     # primeira linha de dados


def ler_csv(caminho):
    with open(caminho, newline="", encoding="utf-8-sig") as fh:
        linhas = list(csv.DictReader(fh, delimiter=";"))
    itens = []
    for i, r in enumerate(linhas, start=1):
        codigo = (r.get("codigo_peca") or "").strip()
        if not codigo:
            continue
        try:
            atual = float((r.get("estoque_atual") or "0").replace(",", "."))
            minimo = float((r.get("estoque_minimo") or "0").replace(",", "."))
        except ValueError:
            sys.exit(f"Linha {i}: estoque_atual/estoque_minimo nao numerico -> {r}")
        itens.append({
            "codigo": codigo,
            "descricao": (r.get("descricao_peca") or "").strip(),
            "atual": int(atual) if atual == int(atual) else atual,
            "minimo": int(minimo) if minimo == int(minimo) else minimo,
        })
    if not itens:
        sys.exit(f"Nenhum item valido em {caminho}")
    return itens


def ordenar(itens):
    """Ruptura primeiro, depois abaixo do minimo, depois no minimo; maior deficit antes."""
    def chave(it):
        if it["atual"] <= 0:
            prio = 1
        elif it["atual"] < it["minimo"]:
            prio = 2
        elif it["atual"] == it["minimo"]:
            prio = 3
        else:
            prio = 4
        deficit = max(it["minimo"] - it["atual"], 0)
        return (prio, -deficit, -it["minimo"], it["codigo"])
    return sorted(itens, key=chave)


def montar(itens, revenda, fator, data_ref):
    wb = Workbook()
    ws = wb.active
    ws.title = "Reposicao Revenda 1"

    ws["A1"] = "PEDIDO DE REPOSIÇÃO — ESTOQUE MÍNIMO"
    ws["A1"].font = TITULO
    ws["A2"] = f"Revenda {revenda}  |  Data de referência: {data_ref:%d/%m/%Y}"
    ws["A2"].font = SUB
    ws["A3"] = "Itens em ruptura, abaixo do mínimo ou exatamente no mínimo (ponto de pedido)."
    ws["A3"].font = SUB

    # --- Premissa editavel -------------------------------------------------
    ws["A5"] = "Fator de reposição:"
    ws["A5"].font = NEGRITO
    ws["C5"] = fator
    ws["C5"].font = AZUL
    ws["C5"].fill = FILL_INPUT
    ws["C5"].border = BORDA
    ws["C5"].number_format = "0.00"
    ws["C5"].alignment = Alignment(horizontal="center")
    ws["C5"].comment = Comment(
        "Multiplicador do estoque mínimo usado para dimensionar o lote de compra.\n"
        "1,00 = comprar exatamente a quantidade do mínimo (repor um lote).\n"
        "2,00 = comprar o dobro do mínimo.\n"
        "Premissa informada pelo usuário — altere APENAS esta célula e a coluna\n"
        "'Qtd. a Comprar' recalcula para todas as peças.",
        "Analise de Estoque",
    )
    ws["D5"] = "← altere aqui (célula amarela): 1,00 = repor um lote igual ao mínimo"
    ws["D5"].font = SUB

    # --- Cabecalho da tabela ----------------------------------------------
    for idx, (titulo, largura) in enumerate(COLUNAS, start=1):
        col = get_column_letter(idx)
        ws.column_dimensions[col].width = largura
        c = ws.cell(row=LINHA_CAB, column=idx, value=titulo)
        c.font = CAB
        c.fill = FILL_CAB
        c.border = BORDA
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[LINHA_CAB].height = 30

    # --- Linhas ------------------------------------------------------------
    for i, it in enumerate(itens):
        r = LINHA_1 + i
        ws.cell(row=r, column=1, value=i + 1).font = PRETO
        ws.cell(row=r, column=2, value=it["codigo"]).font = AZUL
        ws.cell(row=r, column=3, value=it["descricao"] or None).font = AZUL
        ws.cell(row=r, column=4, value=revenda).font = PRETO
        ws.cell(row=r, column=5, value=it["atual"]).font = AZUL
        ws.cell(row=r, column=6, value=it["minimo"]).font = AZUL

        # Deficit = MAX(minimo - atual, 0)
        ws.cell(row=r, column=7, value=f"=MAX(F{r}-E{r},0)").font = PRETO
        # Qtd a comprar = o maior entre o deficit e o lote (minimo * fator)
        ws.cell(row=r, column=8,
                value=f"=MAX(G{r},ROUND(F{r}*$C$5,0))").font = NEGRITO
        # Status
        ws.cell(row=r, column=9, value=(
            f'=IF(E{r}<=0,"Crítico",'
            f'IF(E{r}<F{r},"Abaixo do Mínimo",'
            f'IF(E{r}=F{r},"No Mínimo","Ok")))'
        )).font = PRETO

        for col in range(1, len(COLUNAS) + 1):
            cel = ws.cell(row=r, column=col)
            cel.border = BORDA
            if col in (1, 4, 5, 6, 7, 8):
                cel.alignment = Alignment(horizontal="center")
                if col >= 5:
                    cel.number_format = "#,##0"
            elif col == 9:
                cel.alignment = Alignment(horizontal="center")

    ultima = LINHA_1 + len(itens) - 1

    # --- Total -------------------------------------------------------------
    tr = ultima + 1
    ws.cell(row=tr, column=2, value="TOTAL").font = NEGRITO
    for col in (7, 8):
        L = get_column_letter(col)
        c = ws.cell(row=tr, column=col, value=f"=SUM({L}{LINHA_1}:{L}{ultima})")
        c.font = NEGRITO
        c.number_format = "#,##0"
        c.alignment = Alignment(horizontal="center")
    ws.cell(row=tr, column=3, value=f"=COUNTA(B{LINHA_1}:B{ultima})&\" itens\"").font = NEGRITO
    for col in range(1, len(COLUNAS) + 1):
        ws.cell(row=tr, column=col).fill = FILL_TOTAL
        ws.cell(row=tr, column=col).border = BORDA

    # --- Realce por status --------------------------------------------------
    faixa = f"A{LINHA_1}:I{ultima}"
    for texto, fill in (("Crítico", FILL_CRITICO),
                        ("Abaixo do Mínimo", FILL_ABAIXO),
                        ("No Mínimo", FILL_NO_MINIMO)):
        ws.conditional_formatting.add(
            faixa, FormulaRule(formula=[f'$I{LINHA_1}="{texto}"'], fill=fill, stopIfTrue=False)
        )

    # --- Nota de premissa, visivel ao lado da tabela ------------------------
    nr = tr + 2
    ws.cell(row=nr, column=1, value="PREMISSAS E ORIGEM DOS DADOS").font = NEGRITO
    notas = [
        "Lista de peças e quantidades informada por Fernando de Souza Roquete "
        "(mensagem de 27/08/2026); nenhuma consulta ao Data Lake foi executada.",
        "As peças listadas estão NO estoque mínimo (ponto de pedido): "
        "Estoque Atual = Estoque Mínimo, portanto o Déficit é 0.",
        "Como o déficit é 0, a Qtd. a Comprar vem do lote = Estoque Mínimo x Fator de "
        "Reposição (célula amarela C5, hoje 1,00). Ajuste o fator, ou digite direto na "
        "coluna H, se a política de compra for outra.",
        "Coluna 'Descrição da Peça' vazia de propósito: não foi informada e não deve ser "
        "adivinhada. Preencher pelo cadastro (dim_produto) antes de enviar ao fornecedor.",
        "Células com texto AZUL são dados de entrada; texto PRETO é fórmula calculada.",
    ]
    for j, nota in enumerate(notas):
        c = ws.cell(row=nr + 1 + j, column=1, value=f"• {nota}")
        c.font = Font(name=FONTE, size=9, color="595959")
        c.alignment = Alignment(vertical="top", wrap_text=True)
        ws.merge_cells(start_row=nr + 1 + j, start_column=1, end_row=nr + 1 + j, end_column=9)
        ws.row_dimensions[nr + 1 + j].height = 26

    ws.freeze_panes = f"A{LINHA_1}"
    ws.auto_filter.ref = f"A{LINHA_CAB}:I{ultima}"
    ws.page_setup.orientation = "landscape"
    ws.print_title_rows = f"{LINHA_CAB}:{LINHA_CAB}"

    _aba_legenda(wb)
    return wb


def _aba_legenda(wb):
    ws = wb.create_sheet("Legenda")
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 88
    ws["A1"] = "COMO USAR ESTA PLANILHA"
    ws["A1"].font = TITULO
    linhas = [
        ("Célula amarela (C5)", "Fator de reposição. É a ÚNICA premissa. 1,00 compra um lote igual ao estoque mínimo."),
        ("Texto azul", "Dado de entrada — pode ser digitado/alterado (código, descrição, estoque atual, estoque mínimo)."),
        ("Texto preto", "Fórmula — não digite por cima, recalcula sozinho."),
        ("Déficit", "=MAX(Estoque Mínimo - Estoque Atual; 0). Quanto falta para voltar ao mínimo."),
        ("Qtd. a Comprar", "=MAX(Déficit; ARRED(Estoque Mínimo x Fator; 0)). Nunca menor que o déficit."),
        ("Status", "Crítico = estoque 0 ou negativo · Abaixo do Mínimo · No Mínimo (ponto de pedido) · Ok."),
        ("Cores das linhas", "Vermelho = Crítico · Amarelo = Abaixo do Mínimo · Azul = No Mínimo."),
        ("Exemplo de preenchimento", "Ver linha abaixo: formato esperado de cada coluna."),
    ]
    for i, (a, b) in enumerate(linhas, start=3):
        ws.cell(row=i, column=1, value=a).font = NEGRITO
        c = ws.cell(row=i, column=2, value=b)
        c.font = PRETO
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[i].height = 28

    r = 3 + len(linhas) + 1
    ws.cell(row=r, column=1, value="EXEMPLO (não é dado real)").font = NEGRITO
    exemplo = [("Código da Peça", "5Z0820411E"), ("Descrição da Peça", "Filtro de ar condicionado"),
               ("Estoque Atual", "2"), ("Estoque Mínimo", "2"), ("Resultado esperado", "Status 'No Mínimo', Déficit 0, Qtd. a Comprar 2")]
    for i, (a, b) in enumerate(exemplo, start=r + 1):
        ws.cell(row=i, column=1, value=a).font = PRETO
        ws.cell(row=i, column=2, value=b).font = AZUL
    return ws


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("csv_entrada")
    ap.add_argument("xlsx_saida")
    ap.add_argument("--revenda", type=int, default=1)
    ap.add_argument("--fator", type=float, default=1.0,
                    help="multiplicador do estoque minimo para dimensionar o lote (padrao 1.0)")
    ap.add_argument("--data", default=None, help="AAAA-MM-DD; padrao = hoje")
    args = ap.parse_args()

    data_ref = (dt.date.fromisoformat(args.data) if args.data
                else dt.date.fromtimestamp(int(os.environ.get("SOURCE_DATE_EPOCH", "0")) or
                                           __import__("time").time()))
    itens = ordenar(ler_csv(args.csv_entrada))
    wb = montar(itens, args.revenda, args.fator, data_ref)
    wb.save(args.xlsx_saida)
    print(f"OK: {args.xlsx_saida} ({len(itens)} itens)")


if __name__ == "__main__":
    main()
