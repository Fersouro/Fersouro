#!/usr/bin/env python3
"""
Gera a planilha consolidada "Estratégico VW" a partir da planilha de pedido PAC.

Entrada : Estrategico_PAC_VII_5.xlsx  (aba 'Pedido')
            coluna C = Partnumber (Código da Peça)
            coluna H = Preço Rev. c/ IPI (Promo.)  -> Preço Base
Saída   : Estrategico_VW_Consolidado.xlsx
            aba 'Consolidado' : Código | Estoque (DataLake) | Preço Público (DataLake)
                                | Preço Original (Coluna H) | Preço Estratégico (+40%)
            aba 'DataLake'    : área de colagem do extrato do DataLake
            aba 'Instruções'  : legenda e parâmetro do acréscimo

Uso: python3 gerar_consolidada.py <entrada.xlsx> <saida.xlsx>
"""
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

ORIGEM = sys.argv[1] if len(sys.argv) > 1 else "Estrategico_PAC_VII_5.xlsx"
DESTINO = sys.argv[2] if len(sys.argv) > 2 else "Estrategico_VW_Consolidado.xlsx"

COL_CODIGO = 3   # C - Partnumber
COL_DESCR = 4    # D - Descrição (usada apenas para conferência no DataLake)
COL_H = 8        # H - Preço Rev. c/ IPI (Promo.) = Preço Base

FONTE = "Arial"
AZUL = Font(name=FONTE, size=10, color="0000FF")        # entrada manual
PRETO = Font(name=FONTE, size=10)                        # fórmula / valor
VERDE = Font(name=FONTE, size=10, color="008000")        # link para outra aba
CAB = Font(name=FONTE, size=10, bold=True, color="FFFFFF")
FILL_CAB = PatternFill("solid", fgColor="1F3864")
FILL_INPUT = PatternFill("solid", fgColor="FFFF00")      # preencher aqui
MOEDA = 'R$ #,##0.00;(R$ #,##0.00);-'
BORDA = Border(*[Side(style="thin", color="BFBFBF")] * 4)


def ler_itens(caminho):
    """Extrai (código, descrição, preço_H) das linhas com partnumber real."""
    wb = openpyxl.load_workbook(caminho, data_only=True, read_only=True)
    ws = wb["Pedido"]
    itens = []
    for linha in ws.iter_rows(min_row=5, values_only=True):
        codigo = linha[COL_CODIGO - 1]
        if not isinstance(codigo, str) or not codigo.strip():
            continue  # linhas de preenchimento (0) e linhas vazias do formulário
        preco_h = linha[COL_H - 1]
        if not isinstance(preco_h, (int, float)):
            preco_h = None
        itens.append((codigo.strip(), linha[COL_DESCR - 1], preco_h))
    wb.close()
    return itens


def main():
    itens = ler_itens(ORIGEM)
    n = len(itens)
    ult = n + 1  # última linha de dados (cabeçalho na linha 1)

    wb = openpyxl.Workbook()

    # ------------------------------------------------------------------ #
    # Aba Consolidado — entregável                                        #
    # ------------------------------------------------------------------ #
    ws = wb.active
    ws.title = "Consolidado"
    cabecalhos = [
        "Código da Peça",
        "Estoque (DataLake)",
        "Preço Público (DataLake)",
        "Preço Original (Coluna H)",
        "Preço Estratégico (+40%)",
    ]
    for c, texto in enumerate(cabecalhos, 1):
        cel = ws.cell(row=1, column=c, value=texto)
        cel.font = CAB
        cel.fill = FILL_CAB
        cel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cel.border = BORDA

    ws["D1"].comment = Comment(
        "Coluna H da planilha de origem (aba 'Pedido'): "
        "'Preço Rev. c/ IPI (Promo.)'. Valor copiado da planilha "
        "Estratégico PAC VII enviada pelo usuário.",
        "Consolidação Estratégico VW",
    )
    ws["E1"].comment = Comment(
        "Preço Original x (1 + acréscimo). O acréscimo de 40% está na "
        "célula Instruções!B6 e é referenciado pela fórmula.",
        "Consolidação Estratégico VW",
    )
    ws["B1"].comment = Comment(
        "Buscado no extrato do DataLake (aba 'DataLake') por INDEX/MATCH "
        "usando o Código da Peça como chave.",
        "Consolidação Estratégico VW",
    )

    for i, (codigo, _descr, preco_h) in enumerate(itens):
        r = i + 2
        ws.cell(row=r, column=1, value=codigo).font = PRETO
        # Estoque — INDEX/MATCH no extrato do DataLake
        ws.cell(
            row=r,
            column=2,
            value=(
                f'=IFERROR(IF(INDEX(DataLake!$B$2:$B${ult},'
                f'MATCH($A{r},DataLake!$A$2:$A${ult},0))="","Pendente",'
                f'INDEX(DataLake!$B$2:$B${ult},'
                f'MATCH($A{r},DataLake!$A$2:$A${ult},0))),"Não encontrado")'
            ),
        ).font = VERDE
        # Preço Público — INDEX/MATCH no extrato do DataLake
        ws.cell(
            row=r,
            column=3,
            value=(
                f'=IFERROR(IF(INDEX(DataLake!$C$2:$C${ult},'
                f'MATCH($A{r},DataLake!$A$2:$A${ult},0))="","Pendente",'
                f'INDEX(DataLake!$C$2:$C${ult},'
                f'MATCH($A{r},DataLake!$A$2:$A${ult},0))),"Não encontrado")'
            ),
        ).font = VERDE
        # Preço Original — valor literal vindo da coluna H de origem
        cel_h = ws.cell(row=r, column=4, value=preco_h)
        cel_h.font = AZUL
        cel_h.number_format = MOEDA
        # Preço Estratégico — fórmula sobre o parâmetro de acréscimo
        cel_e = ws.cell(row=r, column=5, value=f"=D{r}*(1+Instruções!$B$6)")
        cel_e.font = PRETO
        cel_e.number_format = MOEDA
        for c in range(1, 6):
            ws.cell(row=r, column=c).border = BORDA
        ws.cell(row=r, column=2).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=3).number_format = MOEDA

    for col, larg in zip("ABCDE", (22, 18, 22, 22, 22)):
        ws.column_dimensions[col].width = larg
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:E{ult}"

    # ------------------------------------------------------------------ #
    # Aba DataLake — área de colagem do extrato                           #
    # ------------------------------------------------------------------ #
    dl = wb.create_sheet("DataLake")
    for c, texto in enumerate(
        ["Código da Peça", "Estoque", "Preço Público"], 1
    ):
        cel = dl.cell(row=1, column=c, value=texto)
        cel.font = CAB
        cel.fill = FILL_CAB
        cel.alignment = Alignment(horizontal="center", vertical="center")
        cel.border = BORDA

    for i, (codigo, _descr, _h) in enumerate(itens):
        r = i + 2
        dl.cell(row=r, column=1, value=codigo).font = PRETO
        for c in (2, 3):
            cel = dl.cell(row=r, column=c)
            cel.font = AZUL
            cel.fill = FILL_INPUT
            cel.border = BORDA
        dl.cell(row=r, column=1).border = BORDA
        dl.cell(row=r, column=3).number_format = MOEDA
        dl.cell(row=r, column=2).alignment = Alignment(horizontal="center")

    nota = dl.cell(
        row=ult + 2,
        column=1,
        value=(
            "Preencha as colunas B (Estoque) e C (Preço Público) com os dados do "
            "extrato do DataLake. Os códigos da coluna A já estão na ordem da "
            "aba 'Consolidado'; a busca é feita por INDEX/MATCH, portanto a ordem "
            "pode ser alterada sem quebrar as fórmulas."
        ),
    )
    nota.font = Font(name=FONTE, size=9, italic=True)
    for col, larg in zip("ABC", (22, 14, 18)):
        dl.column_dimensions[col].width = larg
    dl.freeze_panes = "A2"

    # ------------------------------------------------------------------ #
    # Aba Instruções — legenda e parâmetro                                #
    # ------------------------------------------------------------------ #
    ins = wb.create_sheet("Instruções")
    linhas = [
        ("Estratégico VW — Planilha Consolidada", None),
        ("", None),
        ("Origem dos dados", None),
        ("Planilha de origem", "Estratégico PAC VII (aba 'Pedido')"),
        ("Coluna H de origem", "Preço Rev. c/ IPI (Promo.)"),
        ("Acréscimo Estratégico", 0.40),
        ("Itens processados", n),
        ("", None),
        ("Como completar", None),
        ("1", "Cole o extrato do DataLake na aba 'DataLake' (colunas B e C, células amarelas)."),
        ("2", "A aba 'Consolidado' preenche Estoque e Preço Público automaticamente por INDEX/MATCH."),
        ("3", "'Pendente' = código presente no extrato mas sem valor; 'Não encontrado' = código ausente do extrato."),
        ("", None),
        ("Exemplo de preenchimento na aba DataLake", None),
        ("Código da Peça", "Estoque"),
        ("04C-115-105-E", 37),
        ("", None),
        ("Legenda de cores", None),
        ("Azul", "Valor fixo / entrada manual"),
        ("Preto", "Fórmula calculada nesta aba"),
        ("Verde", "Fórmula que busca dados em outra aba"),
        ("Amarelo", "Célula a ser preenchida pelo usuário"),
    ]
    for i, (a, b) in enumerate(linhas, 1):
        ca = ins.cell(row=i, column=1, value=a)
        ca.font = Font(name=FONTE, size=10, bold=b is None and bool(a))
        if b is not None:
            cb = ins.cell(row=i, column=2, value=b)
            cb.font = AZUL if isinstance(b, (int, float)) else PRETO
    ins["B6"].number_format = "0.0%"
    ins["B6"].fill = FILL_INPUT
    ins["B6"].comment = Comment(
        "Parâmetro do cálculo: Preço Estratégico = Preço Original x (1 + este valor). "
        "Definido em 40% conforme solicitação do usuário. Alterar aqui recalcula "
        "toda a coluna E da aba 'Consolidado'.",
        "Consolidação Estratégico VW",
    )
    ins["A1"].font = Font(name=FONTE, size=13, bold=True)
    ins["B16"].font = AZUL
    ins.column_dimensions["A"].width = 40
    ins.column_dimensions["B"].width = 78

    wb.save(DESTINO)
    print(f"OK: {DESTINO} — {n} itens")


if __name__ == "__main__":
    main()
