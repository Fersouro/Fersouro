#!/usr/bin/env python3
"""
Núcleo da consolidação Estratégico VW.

Lê a planilha de pedido (aba 'Pedido'), lê o extrato do DataLake
(CSV/TSV/XLSX) e cruza os dois pelo Código da Peça, gerando a planilha
consolidada com as cinco colunas do entregável.

Usado tanto pelo servidor web (app.py) quanto pela linha de comando.
"""
from __future__ import annotations

import csv
import io
import re
import unicodedata
from dataclasses import dataclass, field

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# --------------------------------------------------------------------------- #
# Constantes da planilha de origem                                            #
# --------------------------------------------------------------------------- #
ABA_PEDIDO = "Pedido"
PRIMEIRA_LINHA = 5
COL_CODIGO = 3   # C - Partnumber
COL_DESCR = 4    # D - Descrição
COL_H = 8        # H - Preço Rev. c/ IPI (Promo.) = Preço Base

ACRESCIMO_PADRAO = 0.40

CABECALHOS = [
    "Código da Peça",
    "Estoque (DataLake)",
    "Preço Público (DataLake)",
    "Preço Original (Coluna H)",
    "Preço Estratégico (+40%)",
]

# --------------------------------------------------------------------------- #
# Estilos                                                                     #
# --------------------------------------------------------------------------- #
FONTE = "Arial"
AZUL = Font(name=FONTE, size=10, color="0000FF")
PRETO = Font(name=FONTE, size=10)
VERDE = Font(name=FONTE, size=10, color="008000")
CAB = Font(name=FONTE, size=10, bold=True, color="FFFFFF")
FILL_CAB = PatternFill("solid", fgColor="1F3864")
FILL_ALERTA = PatternFill("solid", fgColor="FFF2CC")
MOEDA = 'R$ #,##0.00;(R$ #,##0.00);-'
BORDA = Border(*[Side(style="thin", color="BFBFBF")] * 4)


# --------------------------------------------------------------------------- #
# Normalização e parsing                                                      #
# --------------------------------------------------------------------------- #
def normalizar_codigo(valor) -> str:
    """
    Chave de comparação tolerante: maiúsculas, sem acento e sem separadores.

    '2GX-857-705-  -RAA' e '2GX857705RAA' passam a casar, assim como
    '04c 115 105 e' e '04C-115-105-E'.
    """
    if valor is None:
        return ""
    texto = str(valor).strip()
    if not texto:
        return ""
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    return re.sub(r"[^0-9A-Za-z]", "", texto).upper()


def parse_numero(valor):
    """
    Converte texto numérico em float, aceitando formato pt-BR e en-US.

    Entende 'R$ 1.234,56', '1,234.56', '1234,5', '1.234' e devolve None
    para vazio ou texto não numérico.
    """
    if valor is None:
        return None
    if isinstance(valor, bool):
        return None
    if isinstance(valor, (int, float)):
        return float(valor)

    texto = str(valor).strip()
    if not texto:
        return None
    texto = texto.replace("\xa0", " ")
    texto = re.sub(r"(?i)r\$|un|pç|pcs", "", texto)
    texto = texto.strip()
    negativo = texto.startswith("(") and texto.endswith(")")
    texto = texto.strip("()").strip()
    texto = re.sub(r"[^\d,.\-+]", "", texto)
    if not texto or not re.search(r"\d", texto):
        return None

    tem_ponto, tem_virgula = "." in texto, "," in texto
    if tem_ponto and tem_virgula:
        # o separador decimal é o que aparece por último
        sep = "," if texto.rfind(",") > texto.rfind(".") else "."
        milhar = "." if sep == "," else ","
        texto = texto.replace(milhar, "").replace(sep, ".")
    elif tem_virgula:
        texto = texto.replace(",", ".")
    elif tem_ponto:
        # '1.234' e '1.234.567' são milhar; '1.234' com 1-2 decimais é decimal
        if re.fullmatch(r"[+-]?\d{1,3}(\.\d{3})+", texto):
            texto = texto.replace(".", "")

    try:
        numero = float(texto)
    except ValueError:
        return None
    return -numero if negativo else numero


def _sem_acento_minusculo(texto: str) -> str:
    texto = unicodedata.normalize("NFKD", str(texto or ""))
    return "".join(c for c in texto if not unicodedata.combining(c)).lower()


# --------------------------------------------------------------------------- #
# Leitura da planilha de origem                                               #
# --------------------------------------------------------------------------- #
@dataclass
class Item:
    codigo: str
    descricao: str
    preco_h: float | None
    estoque: float | None = None
    preco_publico: float | None = None
    origem: str = "nao_encontrado"   # exato | normalizado | nao_encontrado | sem_valor


def ler_pedido(fonte) -> list[Item]:
    """Extrai os itens reais da aba 'Pedido' (linhas com partnumber textual)."""
    wb = openpyxl.load_workbook(fonte, data_only=True, read_only=True)
    if ABA_PEDIDO not in wb.sheetnames:
        wb.close()
        raise ValueError(
            f"A planilha enviada não tem a aba '{ABA_PEDIDO}'. "
            f"Abas encontradas: {', '.join(wb.sheetnames)}."
        )
    ws = wb[ABA_PEDIDO]
    itens: list[Item] = []
    for linha in ws.iter_rows(min_row=PRIMEIRA_LINHA, values_only=True):
        if len(linha) < COL_H:
            continue
        codigo = linha[COL_CODIGO - 1]
        # linhas de preenchimento do formulário trazem 0 no partnumber
        if not isinstance(codigo, str) or not codigo.strip():
            continue
        preco = linha[COL_H - 1]
        itens.append(
            Item(
                codigo=codigo.strip(),
                descricao=str(linha[COL_DESCR - 1] or ""),
                preco_h=float(preco) if isinstance(preco, (int, float)) else None,
            )
        )
    wb.close()
    if not itens:
        raise ValueError(
            "Nenhum item encontrado na aba 'Pedido'. Esperado o Partnumber na "
            "coluna C a partir da linha 5."
        )
    return itens


# --------------------------------------------------------------------------- #
# Leitura do extrato do DataLake                                              #
# --------------------------------------------------------------------------- #
@dataclass
class Extrato:
    cabecalhos: list[str]
    linhas: list[list] = field(default_factory=list)


def ler_extrato(dados: bytes, nome_arquivo: str) -> Extrato:
    """Lê o extrato do DataLake a partir de CSV, TSV ou XLSX."""
    nome = (nome_arquivo or "").lower()
    if nome.endswith(".xls"):
        raise ValueError(
            "Formato .xls (Excel 97-2003) não é suportado. Abra o arquivo e "
            "salve como .xlsx ou .csv antes de enviar."
        )
    if nome.endswith((".xlsx", ".xlsm")):
        return _ler_extrato_xlsx(dados)
    return _ler_extrato_texto(dados)


def _ler_extrato_xlsx(dados: bytes) -> Extrato:
    wb = openpyxl.load_workbook(io.BytesIO(dados), data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    linhas = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    return _montar_extrato(linhas)


def _ler_extrato_texto(dados: bytes) -> Extrato:
    texto = None
    for codificacao in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            texto = dados.decode(codificacao)
            break
        except UnicodeDecodeError:
            continue
    if texto is None:
        raise ValueError("Não foi possível decodificar o arquivo enviado.")

    amostra = texto[:8192]
    try:
        dialeto = csv.Sniffer().sniff(amostra, delimiters=";,\t|")
        delimitador = dialeto.delimiter
    except csv.Error:
        # heurística de reserva: o separador mais frequente na primeira linha
        primeira = amostra.splitlines()[0] if amostra.splitlines() else ""
        delimitador = max(";,\t|", key=primeira.count)
        if primeira.count(delimitador) == 0:
            delimitador = ";"

    linhas = [list(r) for r in csv.reader(io.StringIO(texto), delimiter=delimitador)]
    return _montar_extrato(linhas)


def _montar_extrato(linhas: list[list]) -> Extrato:
    linhas = [linha for linha in linhas if any(
        c is not None and str(c).strip() for c in linha)]
    if not linhas:
        raise ValueError("O extrato enviado está vazio.")

    cabecalhos = [
        (str(c).strip() if c is not None and str(c).strip() else f"Coluna {i + 1}")
        for i, c in enumerate(linhas[0])
    ]
    dados = linhas[1:]
    if not dados:
        raise ValueError(
            "O extrato tem apenas a linha de cabeçalho, sem nenhum dado. "
            "Confira se o arquivo exportado do DataLake veio completo."
        )
    largura = len(cabecalhos)
    dados = [linha[:largura] + [None] * (largura - len(linha)) for linha in dados]
    return Extrato(cabecalhos=cabecalhos, linhas=dados)


# --------------------------------------------------------------------------- #
# Detecção automática de colunas                                              #
# --------------------------------------------------------------------------- #
_PISTAS = {
    "codigo": ["codigo da peca", "cod peca", "partnumber", "part number",
               "part_number", "codigo", "cod.", "cod", "peca", "sku",
               "referencia", "item", "material"],
    "estoque": ["estoque", "saldo", "disponivel", "qtd em estoque",
                "quantidade", "qtde", "qtd", "stock", "on hand"],
    "publico": ["preco publico", "preco ao publico", "publico", "pvp",
                "preco de venda", "preco venda", "preco varejo", "varejo",
                "preco", "price", "valor"],
}


def detectar_colunas(extrato: Extrato) -> dict:
    """Sugere qual coluna do extrato é o código, o estoque e o preço público."""
    cabecalhos = [_sem_acento_minusculo(c) for c in extrato.cabecalhos]
    escolhas: dict[str, int | None] = {}
    usadas: set[int] = set()

    for campo in ("codigo", "estoque", "publico"):
        melhor, melhor_nota = None, 0
        for i, cabecalho in enumerate(cabecalhos):
            if i in usadas:
                continue
            nota = 0
            for posicao, pista in enumerate(_PISTAS[campo]):
                if pista == cabecalho:
                    nota = max(nota, 100 - posicao)
                elif pista in cabecalho:
                    nota = max(nota, 60 - posicao)
            if campo == "codigo" and nota == 0 and i == 0:
                nota = 1  # sem pista alguma, a 1ª coluna costuma ser a chave
            if nota > melhor_nota:
                melhor, melhor_nota = i, nota
        escolhas[campo] = melhor
        if melhor is not None:
            usadas.add(melhor)
    return escolhas


# --------------------------------------------------------------------------- #
# Cruzamento                                                                  #
# --------------------------------------------------------------------------- #
def cruzar(itens: list[Item], extrato: Extrato, mapa: dict) -> dict:
    """Preenche estoque e preço público em cada item, casando pelo código."""
    i_cod = mapa.get("codigo")
    if i_cod is None:
        raise ValueError("Informe qual coluna do extrato contém o código da peça.")
    i_est, i_pub = mapa.get("estoque"), mapa.get("publico")
    if i_est is None and i_pub is None:
        raise ValueError(
            "Informe ao menos uma coluna de Estoque ou de Preço Público no extrato."
        )

    por_codigo: dict[str, list] = {}
    por_normal: dict[str, list] = {}
    for linha in extrato.linhas:
        bruto = linha[i_cod]
        if bruto is None or not str(bruto).strip():
            continue
        por_codigo.setdefault(str(bruto).strip(), linha)
        chave = normalizar_codigo(bruto)
        if chave:
            por_normal.setdefault(chave, linha)

    estatisticas = {"exato": 0, "normalizado": 0, "nao_encontrado": 0, "sem_valor": 0}
    for item in itens:
        linha = por_codigo.get(item.codigo)
        origem = "exato"
        if linha is None:
            linha = por_normal.get(normalizar_codigo(item.codigo))
            origem = "normalizado"
        if linha is None:
            item.origem = "nao_encontrado"
            estatisticas["nao_encontrado"] += 1
            continue

        item.estoque = parse_numero(linha[i_est]) if i_est is not None else None
        item.preco_publico = parse_numero(linha[i_pub]) if i_pub is not None else None
        if item.estoque is None and item.preco_publico is None:
            item.origem = "sem_valor"
            estatisticas["sem_valor"] += 1
        else:
            item.origem = origem
            estatisticas[origem] += 1

    estatisticas["total"] = len(itens)
    estatisticas["linhas_extrato"] = len(extrato.linhas)
    return estatisticas


# --------------------------------------------------------------------------- #
# Geração da saída                                                            #
# --------------------------------------------------------------------------- #
def _rotulo(item: Item):
    if item.origem == "nao_encontrado":
        return "Não encontrado"
    return "Pendente"


def gerar_xlsx(itens: list[Item], acrescimo: float, estatisticas: dict) -> bytes:
    """Monta a planilha consolidada; a coluna E é fórmula sobre o parâmetro."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Consolidado"

    cabecalhos = list(CABECALHOS)
    cabecalhos[4] = f"Preço Estratégico (+{acrescimo * 100:.0f}%)"
    for c, texto in enumerate(cabecalhos, 1):
        cel = ws.cell(row=1, column=c, value=texto)
        cel.font = CAB
        cel.fill = FILL_CAB
        cel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cel.border = BORDA

    ws["D1"].comment = Comment(
        "Coluna H da planilha de origem (aba 'Pedido'): "
        "'Preço Rev. c/ IPI (Promo.)'.",
        "Consolidador Estratégico VW",
    )
    ws["E1"].comment = Comment(
        "Preço Original x (1 + acréscimo). O acréscimo está em Parâmetros!B3.",
        "Consolidador Estratégico VW",
    )
    ws["B1"].comment = Comment(
        "Importado do extrato do DataLake, casado pelo Código da Peça.",
        "Consolidador Estratégico VW",
    )

    for i, item in enumerate(itens):
        r = i + 2
        ws.cell(row=r, column=1, value=item.codigo).font = PRETO

        cel_est = ws.cell(row=r, column=2,
                          value=item.estoque if item.estoque is not None else _rotulo(item))
        cel_est.font = AZUL if item.estoque is not None else PRETO
        cel_est.alignment = Alignment(horizontal="center")
        if item.estoque is not None and float(item.estoque).is_integer():
            cel_est.number_format = "#,##0"

        cel_pub = ws.cell(row=r, column=3,
                          value=item.preco_publico if item.preco_publico is not None
                          else _rotulo(item))
        cel_pub.font = AZUL if item.preco_publico is not None else PRETO
        if item.preco_publico is not None:
            cel_pub.number_format = MOEDA

        cel_h = ws.cell(row=r, column=4, value=item.preco_h)
        cel_h.font = AZUL
        cel_h.number_format = MOEDA

        cel_e = ws.cell(row=r, column=5, value=f"=D{r}*(1+Parâmetros!$B$3)")
        cel_e.font = PRETO
        cel_e.number_format = MOEDA

        for c in range(1, 6):
            ws.cell(row=r, column=c).border = BORDA
        if item.origem == "nao_encontrado":
            for c in (1, 2, 3):
                ws.cell(row=r, column=c).fill = FILL_ALERTA

    for col, larg in zip("ABCDE", (22, 18, 22, 22, 22)):
        ws.column_dimensions[col].width = larg
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:E{len(itens) + 1}"

    # ------------------------------ Parâmetros ------------------------------ #
    par = wb.create_sheet("Parâmetros")
    linhas = [
        ("Estratégico VW — Planilha Consolidada", None),
        ("", None),
        ("Acréscimo Estratégico", acrescimo),
        ("Itens processados", estatisticas.get("total", len(itens))),
        ("Linhas lidas do extrato", estatisticas.get("linhas_extrato", 0)),
        ("Casados pelo código exato", estatisticas.get("exato", 0)),
        ("Casados por código normalizado", estatisticas.get("normalizado", 0)),
        ("Encontrados sem valor no extrato", estatisticas.get("sem_valor", 0)),
        ("Não encontrados no extrato", estatisticas.get("nao_encontrado", 0)),
        ("", None),
        ("Origem dos dados", None),
        ("Código da Peça", "Coluna C da aba 'Pedido' (Partnumber)"),
        ("Preço Original", "Coluna H da aba 'Pedido' (Preço Rev. c/ IPI Promo.)"),
        ("Estoque e Preço Público", "Extrato do DataLake enviado pelo usuário"),
        ("", None),
        ("Observações", None),
        ("Preço Estratégico", "Preço Original x (1 + B3). Alterar B3 recalcula a coluna E."),
        ("Normalizado", "Casou ignorando traços, espaços e maiúsculas do código."),
        ("Não encontrado", "Código ausente do extrato; linha destacada em amarelo."),
    ]
    for i, (a, b) in enumerate(linhas, 1):
        ca = par.cell(row=i, column=1, value=a)
        ca.font = Font(name=FONTE, size=10, bold=b is None and bool(a))
        if b is not None:
            cb = par.cell(row=i, column=2, value=b)
            cb.font = AZUL if isinstance(b, (int, float)) else PRETO
    par["A1"].font = Font(name=FONTE, size=13, bold=True)
    par["B3"].number_format = "0.0%"
    par["B3"].comment = Comment(
        "Parâmetro do cálculo. Preço Estratégico = Preço Original x (1 + este valor).",
        "Consolidador Estratégico VW",
    )
    par.column_dimensions["A"].width = 34
    par.column_dimensions["B"].width = 62

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def gerar_csv(itens: list[Item], acrescimo: float) -> bytes:
    """Mesma tabela em CSV pt-BR (separador ';', decimal ',')."""
    saida = io.StringIO()
    escritor = csv.writer(saida, delimiter=";", lineterminator="\r\n")
    cabecalhos = list(CABECALHOS)
    cabecalhos[4] = f"Preço Estratégico (+{acrescimo * 100:.0f}%)"
    escritor.writerow(cabecalhos)

    def br(valor):
        if valor is None:
            return ""
        return f"{valor:,.2f}".replace(",", "_").replace(".", ",").replace("_", ".")

    def br_estoque(valor):
        if valor is None:
            return ""
        if float(valor).is_integer():
            return f"{int(valor):,}".replace(",", ".")
        return br(valor)

    for item in itens:
        rotulo = _rotulo(item)
        estrategico = item.preco_h * (1 + acrescimo) if item.preco_h is not None else None
        escritor.writerow([
            item.codigo,
            br_estoque(item.estoque) if item.estoque is not None else rotulo,
            br(item.preco_publico) if item.preco_publico is not None else rotulo,
            br(item.preco_h),
            br(estrategico),
        ])
    return saida.getvalue().encode("utf-8-sig")


def itens_para_json(itens: list[Item], acrescimo: float) -> list[dict]:
    """Serializa os itens para a pré-visualização na página."""
    return [
        {
            "codigo": item.codigo,
            "descricao": item.descricao,
            "estoque": item.estoque,
            "precoPublico": item.preco_publico,
            "precoOriginal": item.preco_h,
            "precoEstrategico": (item.preco_h * (1 + acrescimo)
                                 if item.preco_h is not None else None),
            "origem": item.origem,
        }
        for item in itens
    ]
